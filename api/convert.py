import io, re, json
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from http.server import BaseHTTPRequestHandler

def clean_cell(s):
    if not s: return ''
    s = str(s)
    s = re.sub(r'(?m)^[A-ZÇĞİÖŞÜ]\n', '', s)
    return re.sub(r'\s+', ' ', s).strip()

def fix_fraction(s):
    if not s: return ''
    s = str(s)
    s = re.sub(r'(?m)^[A-ZÇĞİÖŞÜ]\n', '', s)
    s = re.sub(r'(\d)\n(\d)', r'\1\2', s)
    return re.sub(r'\s+', ' ', s).strip()

def clean_malik(s):
    s = clean_cell(s)
    s = re.sub(r'^[A-ZÇĞİÖŞÜ]\s+(?=\(SN:)', '', s)
    return s.strip()

def extract_name(malik):
    m = re.search(r'\)\s+(.+?)\s+:', malik)
    if m: return m.group(1).strip()
    m2 = re.search(r'\)\s+(.+)', malik)
    return m2.group(1).strip() if m2 else ''

def split_pp(hp):
    c = re.sub(r'\s', '', hp)
    m = re.search(r'(\d+)/(\d+)', c)
    return (m.group(1), m.group(2)) if m else ('', '')

def extract_table(pdf_bytes):
    all_rows, prev = [], None
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for pn in range(2, 8):
            if pn > len(pdf.pages): break
            for tbl in pdf.pages[pn-1].extract_tables():
                for raw in tbl:
                    if not raw or len(raw) < 4: continue
                    c0 = str(raw[0] or '').strip()
                    if (not c0) and prev is not None:
                        c1 = str(raw[1] or '').strip()
                        c3 = str(raw[3] or '').strip()
                        if c1: prev[1] = (prev[1] or '') + ' ' + c1
                        if c3: prev[3] = (prev[3] or '') + c3
                        continue
                    digits = re.sub(r'[^0-9]', '', c0.replace('\n',''))
                    if len(digits) < 6: continue
                    if 'Sistem' in str(raw[1] or '') or 'Malik' in str(raw[1] or ''): continue
                    if prev: all_rows.append(prev)
                    prev = list(raw[:8]) + [''] * max(0, 8 - len(raw[:8]))
                    prev[0] = digits
        if prev: all_rows.append(prev)

    seen, unique = set(), []
    for r in all_rows:
        if r[0] not in seen: seen.add(r[0]); unique.append(r)

    if not unique: raise ValueError('MÜLKİYET BİLGİLERİ tablosu bulunamadı.')

    out = []
    for r in unique:
        malik   = clean_malik(str(r[1] or ''))
        hp      = fix_fraction(str(r[3] or ''))
        pay, pd = split_pp(hp)
        out.append({
            'Sistem No':        r[0],
            'Malik':            malik,
            'Adı-Soyadı':       extract_name(malik),
            'El Birliği No':    clean_cell(str(r[2] or '')) or '-',
            'Hisse Pay/Payda':  hp,
            'Pay':              pay,
            'Payda':            pd,
            'Metrekare':        clean_cell(str(r[4] or '')),
            'Toplam Metrekare': clean_cell(str(r[5] or '')),
            'Edinme Sebebi':    clean_cell(str(r[6] or '')),
            'Terkin Sebebi':    clean_cell(str(r[7] or '')) or '-',
        })
    return out

def make_excel(rows):
    H = ['Sistem No','Malik','Adı-Soyadı','El Birliği No',
         'Hisse Pay/Payda','Pay','Payda','Metrekare',
         'Toplam Metrekare','Edinme Sebebi','Terkin Sebebi']
    NEW = {'Adı-Soyadı','Pay','Payda'}

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'MÜLKİYET BİLGİLERİ'

    ws.merge_cells('A1:K1')
    tc = ws['A1']
    tc.value = 'MÜLKİYET BİLGİLERİ'
    tc.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    tc.fill = PatternFill('solid', start_color='1F3864')
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    for ci, h in enumerate(H, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = PatternFill('solid', start_color='1A5276' if h in NEW else '2E4057')
        c.font = Font(name='Calibri', bold=True, size=10, color='A8DAFF' if h in NEW else 'FFFFFF')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 30

    thin = Side(style='thin', color='D0D0D0')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ri, row in enumerate(rows, 3):
        fill = 'F8F9FA' if ri % 2 else 'FFFFFF'
        for ci, h in enumerate(H, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(h,''))
            c.border = brd
            c.alignment = Alignment(vertical='center', wrap_text=True)
            if h in NEW:
                c.fill = PatternFill('solid', start_color='EBF5FB')
                c.font = Font(name='Calibri', size=9, color='1A5276', bold=True)
            else:
                c.fill = PatternFill('solid', start_color=fill)
                c.font = Font(name='Calibri', size=9)
        ws.row_dimensions[ri].height = 16

    for ci, w in enumerate([14,38,22,13,16,8,10,12,15,28,16], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A3'

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def parse_multipart(content_type, body):
    m = re.search(r'boundary=([^\s;]+)', content_type)
    if not m: return body
    boundary = m.group(1).encode()
    for part in body.split(b'--' + boundary):
        if b'filename' in part and b'.pdf' in part[:300].lower():
            idx = part.find(b'\r\n\r\n')
            if idx != -1: return part[idx+4:].rstrip(b'\r\n--')
    return None

class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self._cors()
        self.end_headers()

    def do_POST(self):
        try:
            length = int(self.headers.get('Content-Length', 0))
            ct     = self.headers.get('Content-Type', '')
            body   = self.rfile.read(length)
            pdf    = parse_multipart(ct, body)
            if not pdf: raise ValueError('PDF verisi alınamadı.')
            rows   = extract_table(pdf)
            xlsx   = make_excel(rows)
            self.send_response(200)
            self._cors()
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="mulkiyet_bilgileri.xlsx"')
            self.send_header('Content-Length', str(len(xlsx)))
            self.end_headers()
            self.wfile.write(xlsx)
        except Exception as e:
            err = json.dumps({'error': str(e)}).encode()
            self.send_response(500)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', str(len(err)))
            self.end_headers()
            self.wfile.write(err)

    def _cors(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')

    def log_message(self, *a): pass
